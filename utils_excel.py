import os
import pandas as pd
from openpyxl import load_workbook

def append_to_excel(file_path, data_dict):
    """
    Appends a dictionary of data as a new row to an Excel file.
    If the file doesn't exist, it creates one.
    
    Args:
        file_path (str): Path to the Excel file.
        data_dict (dict): Dictionary where keys are column headers and values are row data.
    """
    df_new = pd.DataFrame([data_dict])
    
    if os.path.exists(file_path):
        try:
            # Check if file is valid Excel
            pd.read_excel(file_path)
            
            # Load existing workbook to append
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Try to load existing sheet to find the last row
                try:
                    # writer.book = load_workbook(file_path) # Not needed in newer pandas with mode='a'
                    # Assuming data is in the first sheet or default sheet 'Sheet1'
                    if 'Sheet1' in writer.book.sheetnames:
                        start_row = writer.book['Sheet1'].max_row
                        df_new.to_excel(writer, index=False, header=False, startrow=start_row)
                    else:
                        # If Sheet1 doesn't exist (unlikely if created by pandas), write with header
                        df_new.to_excel(writer, index=False)
                except Exception as e:
                    print(f"Error appending to Excel with openpyxl: {e}. Falling back to pandas append (slower).")
                    # Fallback: Read, concat, write (slower but safer for some edge cases)
                    df_old = pd.read_excel(file_path)
                    df_combined = pd.concat([df_old, df_new], ignore_index=True)
                    df_combined.to_excel(file_path, index=False)

        except Exception as e:
            print(f"Error reading existing Excel file: {e}. Creating a new one.")
            df_new.to_excel(file_path, index=False)
    else:
        # File doesn't exist, create it
        ensure_dir_of(file_path)
        df_new.to_excel(file_path, index=False)
    
    print(f"Data appended to {file_path}")

def ensure_dir_of(filepath):
    """Ensures the directory of the file exists."""
    dirpath = os.path.dirname(filepath)
    if dirpath and not os.path.exists(dirpath):
        os.makedirs(dirpath, exist_ok=True)
